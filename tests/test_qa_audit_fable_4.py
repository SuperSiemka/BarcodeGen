# QA audit — independent hostile-tester run TAG=fable_4
# Every test asserts the CURRENT behavior of the app (green suite). Where the
# current behavior is a BUG, the test documents it explicitly (see QA_RAPORT_fable_4.md).
#
# Isolation rules honoured:
#  - registry: app.REG_KEY patched to Software\BarcodeGen_QA_fable_4, key deleted after
#  - all output goes to fresh tempfile.mkdtemp() dirs
#  - no GUI mainloop; only extracted logic is exercised

import io
import math
import os
import time
import hashlib
import tempfile
import inspect
from pathlib import Path

import pytest
from PIL import Image
import zxingcpp

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from generator import (BarcodeGenerator, BarcodeError, safe_filename,
                       _ean13_checksum, _detect_barcode_type)
import generator as generator_module
import app
from app import load_settings, save_settings, DEFAULT_SETTINGS, PARAM_RANGES

# QA registry isolation — never touch the real Software\BarcodeGen
QA_REG_KEY = r"Software\BarcodeGen_QA_fable_4"
app.REG_KEY = QA_REG_KEY

GEN = BarcodeGenerator()
EAN_OK = "5901234123457"


# ── helpers (independent implementations, no reuse of app internals) ──────────

def my_ean13_check(d12: str) -> int:
    """Independent GS1 checksum: 1-indexed odd positions ×1, even ×3."""
    s = sum(int(c) * (3 if (i + 1) % 2 == 0 else 1) for i, c in enumerate(d12))
    return (10 - s % 10) % 10


def expected_module_px(distance: float, scale: float, dpi: int) -> int:
    req = max(0.05, distance) * scale
    px = max(2, round(req * dpi / 25.4))
    if px * 25.4 / dpi < 0.2:                       # 0.2 mm scanner floor
        px = max(2, math.ceil(0.2 * dpi / 25.4))
    return px


def row_has_black(img: Image.Image) -> list:
    g = img.convert("L")
    w, h = g.size
    data = g.tobytes()
    return [min(data[y * w:(y + 1) * w]) < 128 for y in range(h)]


def runs_true(bools):
    out, start = [], None
    for i, b in enumerate(bools):
        if b and start is None:
            start = i
        elif not b and start is not None:
            out.append((start, i - 1)); start = None
    if start is not None:
        out.append((start, len(bools) - 1))
    return out


def black_runs_in_row(img: Image.Image, y: int) -> list:
    g = img.convert("L")
    row = g.tobytes()[y * g.width:(y + 1) * g.width]
    runs, run = [], 0
    for px in row:
        if px < 128:
            run += 1
        elif run:
            runs.append(run); run = 0
    if run:
        runs.append(run)
    return runs


def analyze(img: Image.Image):
    """(bar_row_run, text_row_run, min_bar_px, left_margin_px)"""
    rr = runs_true(row_has_black(img))
    bar_run = rr[0]
    text_run = rr[1] if len(rr) > 1 else None
    y_mid = (bar_run[0] + bar_run[1]) // 2
    bruns = black_runs_in_row(img, y_mid)
    g = img.convert("L")
    row = g.tobytes()[y_mid * g.width:(y_mid + 1) * g.width]
    left = next((i for i, px in enumerate(row) if px < 128), None)
    return bar_run, text_run, min(bruns), left


def text_glyph_gaps(img: Image.Image):
    """Column-profile the text band; return (glyph_run_count, [gap widths])."""
    rr = runs_true(row_has_black(img))
    assert len(rr) >= 2, "no text band found"
    y0, y1 = rr[1]
    g = img.convert("L")
    w = g.width
    data = g.tobytes()
    cols = [False] * w
    for y in range(y0, y1 + 1):
        row = data[y * w:(y + 1) * w]
        for x in range(w):
            if not cols[x] and row[x] < 128:
                cols[x] = True
    cruns = runs_true(cols)
    gaps = [cruns[i + 1][0] - cruns[i][1] - 1 for i in range(len(cruns) - 1)]
    return len(cruns), gaps


def decode_texts(path: Path) -> list:
    return [r.text for r in zxingcpp.read_barcodes(Image.open(path))]


def sha16(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()[:16]


def tmpdir() -> Path:
    return Path(tempfile.mkdtemp(prefix="qa_f4_"))


# ═══ FAZA 1 — poprawność generowania ═══════════════════════════════════════════

MATRIX = [
    # (distance, dpi, scale, height, font_size) — all range extremes covered
    (0.05, 72,   0.5, 9,   2.2),
    (0.05, 300,  1,   9,   2.2),
    (0.05, 1200, 0.5, 9,   2.2),   # 0.2 mm floor kicks in → 10 px
    (0.2,  300,  3,   9,   2.2),
    (0.33, 72,   1,   9,   2.2),
    (0.33, 300,  1,   1,   0.5),   # min height, min font
    (0.33, 300,  1,   150, 15),    # max height, max font
    (1,    300,  0.5, 9,   8),
    (1,    1200, 1,   9,   2.2),
    (5,    72,   3,   9,   2.2),
    (5,    300,  1,   9,   2.2),
    (10,   300,  1,   9,   2.2),
]


@pytest.mark.parametrize("code", [EAN_OK, "ABC-123"])
def test_f1_matrix(code):
    out = tmpdir()
    hashes = {}
    for i, (d, dpi, sc, h, fs) in enumerate(MATRIX):
        sub = out / f"c{i}"
        sub.mkdir()
        p = GEN.generate(code, sub, height=h, distance=d, font_size=fs, dpi=dpi, scale=sc)
        assert p.exists() and p.stat().st_size > 0, (d, dpi, sc, h, fs)
        assert code in decode_texts(p), f"decode failed for {(d, dpi, sc, h, fs)}"
        _, _, minbar, _ = analyze(Image.open(p))
        exp = expected_module_px(d, sc, dpi)
        assert minbar == exp, f"min bar {minbar}px != expected {exp}px for {(d, dpi, sc)}"
        hashes.setdefault(sha16(p), []).append((d, dpi, sc, h, fs))
    # distinct physical parameters must give distinct files (this sample avoids
    # quantization-equal combos, so every hash must be unique here)
    dupes = {h: v for h, v in hashes.items() if len(v) > 1}
    assert not dupes, f"unexpected identical outputs: {dupes}"


def test_f1_quantization_collapses_distances_at_72dpi():
    """DOCUMENTED behavior: at 72 DPI the 2-px module floor collapses
    distance 0.05 / 0.2 / 0.33 mm into byte-identical files (all → 2 px)."""
    out = tmpdir()
    hs = set()
    for d in (0.05, 0.2, 0.33):
        sub = out / str(d).replace(".", "_")
        sub.mkdir()
        p = GEN.generate(EAN_OK, sub, distance=d, dpi=72, scale=1)
        hs.add(sha16(p))
    assert len(hs) == 1, "expected identical files due to px quantization"


def test_f1_font_absolute_pixel_identical():
    """font_size=3 mm @300 DPI → glyph ink height measured from the IMAGE is
    identical to the pixel across configs, and ≈ 3*300/25.4 = 35.4 px ± 2."""
    out = tmpdir()
    cfgs = [dict(height=1, distance=0.33, scale=1), dict(height=9, distance=0.33, scale=1),
            dict(height=150, distance=0.33, scale=1), dict(height=9, distance=5, scale=1),
            dict(height=9, distance=0.33, scale=3)]
    measured = set()
    for i, c in enumerate(cfgs):
        sub = out / f"f{i}"
        sub.mkdir()
        p = GEN.generate(EAN_OK, sub, font_size=3.0, dpi=300, **c)
        _, text_run, _, _ = analyze(Image.open(p))
        measured.add(text_run[1] - text_run[0] + 1)
    assert len(measured) == 1, f"text height varies across configs: {measured}"
    h = measured.pop()
    assert abs(h - 3.0 * 300 / 25.4) <= 2, f"text {h}px vs target 35.4px"


def test_f1_font_mm_calibration_across_dpi():
    out = tmpdir()
    for dpi in (72, 300, 1200):
        sub = out / str(dpi)
        sub.mkdir()
        p = GEN.generate(EAN_OK, sub, font_size=3.0, dpi=dpi)
        _, text_run, _, _ = analyze(Image.open(p))
        h_px = text_run[1] - text_run[0] + 1
        target = max(6, round(3.0 * dpi / 25.4))
        assert abs(h_px - target) <= 2, f"dpi={dpi}: {h_px}px vs {target}px"


def test_f1_ean13_20_good_20_bad():
    import random
    random.seed(1234)
    out = tmpdir()
    for n in range(20):
        d12 = "".join(random.choice("0123456789") for _ in range(12))
        code = d12 + str(my_ean13_check(d12))
        sub = out / f"g{n}"
        sub.mkdir()
        p = GEN.generate(code, sub)
        assert code in decode_texts(p), code
    for n in range(20):
        d12 = "".join(random.choice("0123456789") for _ in range(12))
        good = my_ean13_check(d12)
        wrong = (good + 1 + n % 8) % 10
        assert wrong != good
        with pytest.raises(BarcodeError) as ei:
            GEN.generate(d12 + str(wrong), out)
        assert f"oczekiwano cyfry kontrolnej {good}" in str(ei.value)


def test_f1_code128_boundaries():
    out = tmpdir()
    for name, code in [("space", "A B"), ("tilde", "A~B"), ("one", "X"),
                       ("eighty", "AB" * 40), ("multispace", "A   B")]:
        sub = out / name
        sub.mkdir()
        p = GEN.generate(code, sub)
        assert decode_texts(p) == [code], name
    for bad in ("AB\x1fCD", "AB\x7fCD"):
        with pytest.raises(BarcodeError, match="niedozwolone znaki"):
            GEN.generate(bad, out)
    with pytest.raises(BarcodeError, match="Pusty kod"):
        GEN.generate("   ", out)   # strip() makes an all-space code empty


# ═══ FAZA 2 — pliki i system ══════════════════════════════════════════════════

def test_f2_windows_reserved_names():
    out = tmpdir()
    expected = {"CON": "_CON.png", "PRN": "_PRN.png", "NUL": "_NUL.png",
                "com5": "_com5.png", "lpt9": "_lpt9.png"}
    for code, fname in expected.items():
        p = GEN.generate(code, out)
        assert p.name == fname
        assert p.exists() and p.stat().st_size > 0, code


def test_f2_name_collision_logic():
    assert safe_filename("AB/CD") == "AB_CD"
    assert safe_filename("AB?CD") == "AB_CD"
    # replica of the pre-generation collision detection in app._on_generate
    codes = ["AB/CD", "AB?CD", "OK-1"]
    by_name = {}
    for c in codes:
        by_name.setdefault(safe_filename(c), []).append(c)
    collisions = [", ".join(v) for v in by_name.values() if len(v) > 1]
    assert collisions == ["AB/CD, AB?CD"]


def test_f2_overwrite_updates_mtime():
    out = tmpdir()
    p1 = GEN.generate("MTIME-XY", out)
    t1 = os.stat(p1).st_mtime_ns
    time.sleep(0.05)
    p2 = GEN.generate("MTIME-XY", out)
    assert p2 == p1
    assert os.stat(p2).st_mtime_ns > t1


def test_f2_200_char_code_long_path():
    """383-char total path succeeded on this machine (LongPathsEnabled=1 +
    Python long-path manifest). On stock Win10 (LongPathsEnabled=0) the same
    call would raise — flagged in report as environment-dependent."""
    deep = tmpdir() / ("sub_" + "x" * 120)
    deep.mkdir(parents=True)
    code = "L" * 200
    assert len(str(deep / (code + ".png"))) > 260
    p = GEN.generate(code, deep)
    assert p.exists() and p.stat().st_size > 0


def test_f2_BUG_device_name_with_extension_crashes():
    """FIXED (Ś2): safe_filename now guards the stem up to the FIRST dot, so
    'con.x' → '_con.x' and the file is written cleanly instead of crashing on
    the reserved CON device. Verifies the fix."""
    out = tmpdir()
    for code in ("con.x", "nul.x", "aux.data", "com1.v2"):
        assert safe_filename(code).startswith("_"), code
        p = GEN.generate(code, out)          # no longer raises
        assert p.exists() and p.stat().st_size > 0
        assert p.name == "_" + code + ".png"


# ═══ FAZA 3 — rejestr i Excel ═════════════════════════════════════════════════

class _QaRegistry:
    def __enter__(self):
        self._wipe()
        return self

    def put(self, **vals):
        import winreg
        self._wipe()
        key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, QA_REG_KEY)
        for k, v in vals.items():
            winreg.SetValueEx(key, k, 0, winreg.REG_SZ, str(v))
        winreg.CloseKey(key)

    def _wipe(self):
        import winreg
        try:
            winreg.DeleteKey(winreg.HKEY_CURRENT_USER, QA_REG_KEY)
        except FileNotFoundError:
            pass

    def __exit__(self, *a):
        self._wipe()


def test_f3_registry_validation_and_migrations():
    assert app.REG_KEY == QA_REG_KEY   # isolation guard
    with _QaRegistry() as reg:
        cases = [
            (dict(language="DE"), "language", "PL"),
            (dict(theme="neon"), "theme", "dark"),
            (dict(dpi="99999"), "dpi", 1200),
            # legacy migration: 12 * 1.693 = 20.316 → round 20.3 → clamp 15.0
            (dict(font_size="12"), "font_size", 15.0),
            (dict(font_size="12", font_unit="mm"), "font_size", 12.0),
            (dict(distance="0.15"), "distance", 0.33),
            (dict(scale="-5"), "scale", 0.5),
            (dict(height="1e6"), "height", 150.0),
            # FIXED (N8): non-finite "nan" now falls back to the default (1.0)
            # instead of silently clamping to the maximum via the min/max chain.
            (dict(scale="nan"), "scale", 1.0),
        ]
        for vals, key, expected in cases:
            reg.put(**vals)
            s = load_settings()           # must not raise
            assert s[key] == expected, f"{vals} -> {key}={s[key]}, expected {expected}"
        # round-trip through save_settings
        reg.put()
        s = dict(DEFAULT_SETTINGS)
        s["dpi"] = 600
        save_settings(s)
        assert load_settings()["dpi"] == 600


def test_f3_excel_import_logic_and_float_ean():
    import openpyxl, datetime
    out = tmpdir()
    xp = out / "t.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 123
    ws["A2"] = 5907925017654.0        # float with .0 — comes back as int via openpyxl
    ws["A3"] = "ABC-123"
    ws["A4"] = "=A1*2"                # no cached value → None under data_only
    ws["A6"] = datetime.datetime(2026, 1, 5)
    ws["A7"] = "   "                  # whitespace-only → skipped
    for i in range(8, 508):
        ws.cell(row=i, column=1, value=f"C{i}")
    wb.save(xp)
    # exact replica of app._import_excel parsing
    wb2 = openpyxl.load_workbook(xp, read_only=True, data_only=True)
    ws2 = wb2.worksheets[0]
    codes = [str(row[0]).strip() for row in ws2.iter_rows(min_row=1, values_only=True)
             if row and row[0] is not None and str(row[0]).strip()]
    wb2.close()
    assert codes[:4] == ["123", "5907925017654", "ABC-123", "2026-01-05 00:00:00"]
    assert len(codes) == 504
    # BUG (ŚREDNI): a 13-digit EAN carrying an Excel float artifact ".0" is
    # silently generated as Code 128 — no warning, wrong symbology.
    assert my_ean13_check("590792501765") == 4       # base IS a valid EAN-13
    fl = "5907925017654.0"
    assert _detect_barcode_type(fl) == "code128"
    sub = out / "gen"
    sub.mkdir()
    p = GEN.generate(fl, sub)
    res = zxingcpp.read_barcodes(Image.open(p))
    assert [(r.text, str(r.format)) for r in res] == [(fl, "Code 128")]


# ═══ FAZA 4 — logika GUI (bez mainloop) ═══════════════════════════════════════

def test_f4_settings_range_validation_logic():
    # replica of app._open_settings._save guard
    def rejected(key, val):
        lo, hi = PARAM_RANGES[key]
        return not (lo <= val <= hi)
    assert rejected("dpi", 10000)
    assert rejected("dpi", 0)
    assert rejected("height", float("nan"))      # NaN comparisons are False → rejected
    assert rejected("distance", 11)
    assert not rejected("dpi", 300)


def test_f4_duplicate_codes_logic():
    codes = ["X", "X", "X", "Y"]
    seen, dupes, unique = set(), [], []
    for c in codes:
        if c in seen:
            if c not in dupes:
                dupes.append(c)
        else:
            seen.add(c)
            unique.append(c)
    assert dupes == ["X"]
    assert unique == ["X", "Y"]


def test_f4_batch_with_two_bad_checksums():
    """Replica of the worker() loop: 10 codes, 2 broken checksums →
    8 PNG files, 2 errors naming the expected check digit."""
    out = tmpdir()
    good = []
    for i in range(8):
        d12 = f"5901234123{i:02d}"
        good.append(d12 + str(my_ean13_check(d12)))
    bad = []
    for d12 in ("590123412345", "123456789012"):
        w = (my_ean13_check(d12) + 5) % 10
        bad.append(d12 + str(w))
    codes = good[:4] + bad[:1] + good[4:7] + bad[1:] + good[7:]
    errors = []
    for code in codes:                      # same try/except shape as worker()
        try:
            GEN.generate(code, out)
        except Exception as e:
            errors.append((code, str(e)))
    assert len(list(out.glob("*.png"))) == 8
    assert len(errors) == 2
    assert all("oczekiwano cyfry kontrolnej" in msg for _, msg in errors)
    assert [c for c, _ in errors] == bad


def test_f4_output_dir_resolution_logic():
    # empty setting → app_dir()/output (project dir when not frozen)
    assert app.app_dir() == Path(app.__file__).parent
    # non-existent nested path → created by mkdir(parents=True)
    base = tmpdir() / "a" / "b" / "c"
    base.mkdir(parents=True, exist_ok=True)
    assert base.is_dir()
    # a FILE at the target path → OSError family (caught in _on_generate)
    f = tmpdir() / "afile"
    f.write_text("x")
    with pytest.raises(OSError):
        f.mkdir(parents=True, exist_ok=True)
    # bad drive → OSError family too
    with pytest.raises(OSError):
        Path(r"Q:\qa_f4\nope").mkdir(parents=True, exist_ok=True)


def test_f4_scale_entry_parse_logic():
    # replica of app._on_scale_entry
    def parse(raw):
        try:
            val = float(raw.replace(",", "."))
            return max(0.5, min(3.0, round(val, 1)))
        except ValueError:
            return "restored"
    assert parse("1,5") == 1.5
    assert parse("abc") == "restored"
    assert parse("0.1") == 0.5
    assert parse("99") == 3.0
    # QUIRK: float() accepts these; NaN slips through min/max to become 3.0
    assert parse("nan") == 3.0
    assert parse("inf") == 3.0
    assert parse("1e1") == 3.0


# ═══ FAZA 4b — odstępy tekstu + wymagania klienta ═════════════════════════════

def test_f4b_ean13_gap_after_7th_and_tracking():
    out = tmpdir()
    p = GEN.generate(EAN_OK, out, font_size=3.0, dpi=300)
    nglyphs, gaps = text_glyph_gaps(Image.open(p))
    assert nglyphs == 13
    assert len(gaps) == 12
    gap7 = gaps[6]                       # gap AFTER the 7th character
    others = gaps[:6] + gaps[7:]
    assert gap7 > 3 * max(others), f"gap7={gap7}px vs others={others}"
    # tracking keeps the other gaps in one modest band (glyph-shape dependent)
    assert max(others) - min(others) <= 10, others


def test_f4b_code128_short_no_7th_char_ok():
    out = tmpdir()
    p = GEN.generate("ABC", out, font_size=3.0, dpi=300)
    assert decode_texts(p) == ["ABC"]
    nglyphs, gaps = text_glyph_gaps(Image.open(p))
    assert nglyphs == 3
    assert max(gaps) - min(gaps) <= 6    # uniform — no stray big gap


def test_f4b_code128_long_gets_the_gap_too():
    """DOCUMENTED design decision: the 'after 7th char' gap is applied to ANY
    code, so a 10-char Code128 shows a 2.2-digit hole mid-text."""
    out = tmpdir()
    p = GEN.generate("ABCDEFGH12", out, font_size=3.0, dpi=300)
    nglyphs, gaps = text_glyph_gaps(Image.open(p))
    assert nglyphs == 10
    assert gaps[6] > 3 * max(gaps[:6] + gaps[7:])


# ═══ FAZA 5 — polowanie swobodne ══════════════════════════════════════════════

def test_f5_BUG_superscript_digit_leaks_valueerror():
    """FIXED (N1): '²' is a non-ASCII digit, so _detect_barcode_type now routes
    '590123412345²' to Code128, where the >126 guard rejects it with a clean
    BarcodeError instead of a raw ValueError from int()."""
    code = "590123412345²"
    assert _detect_barcode_type(code) == "code128"   # no longer misread as EAN-13
    with pytest.raises(BarcodeError):
        GEN.generate(code, tmpdir())


def test_f5_BUG_arabic_indic_digits_content_mismatch():
    """FIXED (Ś3): Arabic-Indic '٥٩٠١٢٣٤١٢٣٤٥٧' is non-ASCII, so it no longer
    passes as EAN-13 (which would encode DIFFERENT Western digits than typed).
    It now routes to Code128 and is rejected cleanly as >126 characters."""
    ar = "٥٩٠١٢٣٤١٢٣٤٥٧"
    assert _detect_barcode_type(ar) == "code128"
    with pytest.raises(BarcodeError):
        GEN.generate(ar, tmpdir())


def test_f5_leading_zero_ean_roundtrip():
    out = tmpdir()
    for d12 in ("000000000000", "000123456789"):
        code = d12 + str(my_ean13_check(d12))
        sub = out / d12
        sub.mkdir()
        p = GEN.generate(code, sub)
        assert decode_texts(p) == [code]     # zxing keeps the 13-digit EAN form


def test_f5_BUG_quiet_zone_cropped_to_one_module():
    """BUG (ŚREDNI): the generator crops the 6.5 mm quiet zone down to
    max(4px, 1 module). EAN-13 spec requires ≥11 modules on the left —
    printed as-is (cut on the picture edge) codes may not scan."""
    out = tmpdir()
    p = GEN.generate(EAN_OK, out, distance=0.33, dpi=300)
    _, _, module_px, left_margin = analyze(Image.open(p))
    assert left_margin <= 2 * module_px, (left_margin, module_px)
    assert left_margin < 11 * module_px      # violates EAN quiet-zone minimum


def test_f5_dead_code_and_unused_import():
    """FIXED (N6): the dead _count_data_modules (+_PROBE_* consts) and the unused
    'import time' were removed from generator.py."""
    src = inspect.getsource(generator_module)
    assert "_count_data_modules" not in src
    assert "_PROBE_MW_MM" not in src and "_PROBE_QZ_MM" not in src
    assert "import time" not in src


def test_f5_generator_errors_polish_only():
    """NISKI: BarcodeError texts are hard-coded Polish — shown verbatim in the
    EN UI ([ERR] log lines / error summary)."""
    with pytest.raises(BarcodeError) as ei:
        GEN.generate("5901234123450", tmpdir())
    assert "Błędna suma kontrolna" in str(ei.value)
    with pytest.raises(BarcodeError) as ei2:
        GEN.generate("", tmpdir())
    assert "Pusty kod" in str(ei2.value)


def test_f5_BUG_param_ranges_allow_gigapixel_image():
    """BUG (KRYTYCZNY): every value below is INSIDE PARAM_RANGES, yet the
    resulting bar image is ~2.9 GIGApixels — measured live run: 118 s of CPU,
    then PIL DecompressionBombError (limit 178,956,970 px). The comment above
    PARAM_RANGES claims these ranges prevent exactly this. Analytic assertion
    here (the live 118 s run is documented in the report)."""
    d, sc, dpi, h = 10.0, 3.0, 1200, 150.0     # all within PARAM_RANGES + scale<=3
    lo, hi = PARAM_RANGES["distance"]; assert lo <= d <= hi
    lo, hi = PARAM_RANGES["dpi"];      assert lo <= dpi <= hi
    lo, hi = PARAM_RANGES["height"];   assert lo <= h <= hi
    module_px = expected_module_px(d, sc, dpi)
    width_px = module_px * 113 + 2 * round(6.5 * dpi / 25.4)   # modules + quiet zone
    height_px = round(h * sc * dpi / 25.4)
    assert width_px * height_px > 2 * Image.MAX_IMAGE_PIXELS   # → DecompressionBombError


def test_f5_registry_key_cleanup():
    """Runs last alphabetically-independent — ensure the QA registry subkey is
    gone and the REAL Software\\BarcodeGen key was never our target."""
    import winreg
    assert app.REG_KEY == QA_REG_KEY
    try:
        winreg.DeleteKey(winreg.HKEY_CURRENT_USER, QA_REG_KEY)
    except FileNotFoundError:
        pass
    with pytest.raises(FileNotFoundError):
        winreg.OpenKey(winreg.HKEY_CURRENT_USER, QA_REG_KEY)
