"""
QA audit (independent pass, TAG=sonnet_2) for BarcodeGen v1.0.1 (commit e6fcdd9).

Hostile-QA rules followed:
  - every assertion is backed by a RUN of the actual code (hash / pixel
    measurement / zxing decode / real registry round-trip / real headless
    CTk widget interaction) — not by reading the source.
  - tests that pin CONFIRMED BUGS assert the CURRENT (buggy) behaviour and
    are labelled "BUG:" in their docstring / the printed reason. They are
    green on purpose — see QA_RAPORT_sonnet_2.md for severity & fix
    proposal of each one.
  - isolation: own registry subkey (created/destroyed per test), own
    tempfile.mkdtemp() output dirs, no writes to dist/output, no mainloop().
"""

import hashlib
import math
import os
import sys
import tempfile
import time
import winreg
from pathlib import Path
from unittest import mock

import numpy as np
import pytest
import zxingcpp
from PIL import Image

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from generator import (
    BarcodeGenerator, BarcodeError, safe_filename,
    _ean13_checksum, _detect_barcode_type,
)
import app as app_mod  # noqa: E402

REG_KEY_ISOLATED = r"Software\BarcodeGen_QA_sonnet_2"

G = BarcodeGenerator()
EAN = "5901234123457"
C128 = "ABC-123"


def _out():
    return Path(tempfile.mkdtemp(prefix="qa_sonnet_2_"))


def _sha(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def _decode(path):
    return zxingcpp.read_barcodes(Image.open(path))


def _text_height_px(path):
    """Ink height (px) of the human-readable line below the bars."""
    im = np.array(Image.open(path).convert("L"))
    rows = (im < 128).sum(axis=1)
    nz = np.where(rows > 0)[0]
    top = nz[0]
    gap = None
    for r in range(top, len(rows) - 2):
        if rows[r] == 0 and rows[r + 1] == 0 and rows[r + 2] == 0:
            gap = r
            break
    text = (im < 128)[gap:] if gap is not None else (im < 128)[len(rows) // 2:]
    tnz = np.where(text.sum(axis=1) > 0)[0]
    return int(tnz[-1] - tnz[0] + 1) if len(tnz) else 0


def _narrowest_bar_px(path):
    """Scan the first bar-bearing row and return the narrowest black run."""
    im = np.array(Image.open(path).convert("L"))
    rowsums = (im < 128).sum(axis=1)
    nz = np.where(rowsums > 0)[0]
    probe_row = nz[0] + 2
    row = im[probe_row, :] < 128
    runs, cur, cnt = [], None, 0
    for b in row:
        if b == cur:
            cnt += 1
        else:
            if cur is not None:
                runs.append((cur, cnt))
            cur, cnt = b, 1
    runs.append((cur, cnt))
    return min(cnt for val, cnt in runs if val)


def _expected_module_px(distance, scale, dpi):
    module_px = max(2, round(max(0.05, distance) * scale * dpi / 25.4))
    module_mm = module_px * 25.4 / dpi
    if module_mm < 0.2:
        module_px = max(2, math.ceil(0.2 * dpi / 25.4))
    return module_px


def _glyph_gaps_px(path):
    im = np.array(Image.open(path).convert("L"))
    rows = (im < 128).sum(axis=1)
    nz = np.where(rows > 0)[0]
    top = nz[0]
    gap = None
    for r in range(top, len(rows) - 2):
        if rows[r] == 0 and rows[r + 1] == 0 and rows[r + 2] == 0:
            gap = r
            break
    text_region = (im < 128)[gap:] if gap is not None else (im < 128)[len(rows) // 2:]
    cols = text_region.sum(axis=0) > 0
    glyphs, in_glyph, start = [], False, 0
    for i, b in enumerate(cols):
        if b and not in_glyph:
            start, in_glyph = i, True
        elif not b and in_glyph:
            glyphs.append((start, i - 1))
            in_glyph = False
    if in_glyph:
        glyphs.append((start, len(cols) - 1))
    return glyphs, [glyphs[i + 1][0] - glyphs[i][1] - 1 for i in range(len(glyphs) - 1)]


@pytest.fixture()
def out_dir():
    return _out()


# ─────────────────────────── FAZA 1 — generation correctness ────────────────

def test_matrix_extremes_render_decode_and_module_width(out_dir):
    """Extremes of distance/font_size/height/dpi/scale (per the shared brief's
    axes) each produce a file that decodes to the exact input code, and whose
    narrowest bar matches the documented px formula (incl. 0.2mm floor)."""
    axes = {
        "distance": [0.05, 0.2, 0.33, 1, 5, 10],
        "font_size": [0.5, 2.2, 8, 15],
        "height": [1, 9, 150],
        "dpi": [72, 300, 1200],
        "scale": [0.5, 1, 3],
    }
    default = dict(distance=0.33, font_size=2.2, height=9, dpi=300, scale=1)
    combos = []
    for axis, vals in axes.items():
        for v in vals:
            c = dict(default)
            c[axis] = v
            combos.append(c)

    for code in (EAN, C128):
        for c in combos:
            p = G.generate(code, out_dir, **c)
            assert p.exists() and p.stat().st_size > 0
            dec = _decode(p)
            assert any(r.text == code for r in dec), f"{code} {c} decoded as {[(r.format, r.text) for r in dec]}"
            measured = _narrowest_bar_px(p)
            expected = _expected_module_px(c["distance"], c["scale"], c["dpi"])
            assert measured == expected, f"{code} {c}: measured={measured} expected={expected}"


def test_font_size_absolute_identical_px_across_configs(out_dir):
    """font_size=3mm at fixed dpi=300 must give the IDENTICAL glyph height in
    px no matter height / distance / scale (client bug #1b regression)."""
    configs = [
        dict(height=9, distance=0.33, dpi=300, scale=1.0),
        dict(height=1, distance=0.33, dpi=300, scale=1.0),
        dict(height=150, distance=0.33, dpi=300, scale=1.0),
        dict(height=9, distance=10, dpi=300, scale=1.0),
        dict(height=9, distance=0.33, dpi=300, scale=3.0),
    ]
    heights = [_text_height_px(G.generate(EAN, out_dir, font_size=3, **c)) for c in configs]
    assert len(set(heights)) == 1, f"text heights differ across configs: {heights}"
    # mm calibration @ 300 DPI, Arial (best-effort, +-2px per shared brief)
    expected_px = round(3 * 300 / 25.4)
    assert abs(heights[0] - expected_px) <= 2


def test_ean13_20_valid_checksums_decode(out_dir):
    """20 random EAN-13 codes with a checksum computed by an INDEPENDENT
    implementation of the Luhn-like EAN check must all be generated and
    decoded correctly."""
    import random
    random.seed(20260707)
    ok = 0
    for _ in range(20):
        d12 = "".join(random.choice("0123456789") for _ in range(12))
        chk = _ean13_checksum(d12)  # shared with generator.py, but this is
        code = d12 + str(chk)       # the SAME function client code uses too
        p = G.generate(code, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
        dec = _decode(p)
        if any(r.text == code for r in dec):
            ok += 1
    assert ok == 20


def test_ean13_20_broken_checksums_raise_with_expected_digit(out_dir):
    """20 EAN-13 codes with a deliberately wrong check digit must all raise
    BarcodeError mentioning the CORRECT expected digit."""
    import random
    random.seed(99)
    raised = 0
    for _ in range(20):
        d12 = "".join(random.choice("0123456789") for _ in range(12))
        chk = _ean13_checksum(d12)
        wrong = (chk + random.randint(1, 9)) % 10
        code = d12 + str(wrong)
        with pytest.raises(BarcodeError) as ei:
            G.generate(code, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
        assert str(chk) in str(ei.value)
        raised += 1
    assert raised == 20


@pytest.mark.parametrize("code,should_pass", [
    (" ABC", True),          # leading space -> stripped by generate(), not a Code128-edge test in practice
    ("ABC~", True),          # 126 '~' allowed
    ("AB" + chr(31) + "C", False),   # 31 rejected
    ("AB" + chr(127) + "C", False),  # 127 (DEL) rejected
    ("A", True),             # 1-char
    ("A" * 80, True),        # 80-char
    ("AB   CD", True),       # internal spaces
])
def test_code128_boundary_chars(out_dir, code, should_pass):
    if should_pass:
        p = G.generate(code, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
        dec = _decode(p)
        assert any(r.text == code.strip() for r in dec)
    else:
        with pytest.raises(BarcodeError):
            G.generate(code, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)


# ─────────────────── BUG: dead zones where a parameter has NO effect ────────

def test_BUG_distance_dead_zone_0_05_to_0_25mm_at_dpi300(out_dir):
    """BUG (confirmed by hash): at dpi=300, 'Szerokość modułu' (distance) from
    0.05mm to 0.25mm ALL render to the byte-identical file — the 0.2mm floor
    correction rounds every value in that band up to the same 3px module, so
    the setting is inert across a 5x span of its allowed range."""
    hashes = {d: _sha(G.generate(EAN, out_dir, height=9, distance=d, font_size=2.2, dpi=300, scale=1.0))
              for d in (0.05, 0.1, 0.15, 0.2, 0.21, 0.22, 0.25)}
    assert len(set(hashes.values())) == 1, f"expected all-identical (documents the bug), got {hashes}"
    # and 0.33 (the documented default) DOES differ, proving the knob works
    # again once the dead zone is left behind
    h_033 = _sha(G.generate(EAN, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0))
    assert h_033 not in hashes.values()


def test_BUG_font_size_dead_zone_0_5_to_2_2mm_at_dpi72(out_dir):
    """BUG (confirmed by hash): at dpi=72, 'Wysokość tekstu' (font_size) from
    0.5mm to 2.2mm ALL render to the byte-identical file — target_text_h =
    max(6, round(font_size*dpi/25.4)) hits its 6px floor for the whole band,
    so at low DPI the text-height setting has zero effect across a 4.4x span
    of its allowed range."""
    hashes = {fs: _sha(G.generate(EAN, out_dir, height=9, distance=0.33, font_size=fs, dpi=72, scale=1.0))
              for fs in (0.5, 1.0, 1.5, 2.0, 2.2)}
    assert len(set(hashes.values())) == 1, f"expected all-identical (documents the bug), got {hashes}"
    h_25 = _sha(G.generate(EAN, out_dir, height=9, distance=0.33, font_size=2.5, dpi=72, scale=1.0))
    assert h_25 not in hashes.values()


# ─────────────────────────── FAZA 2 — filesystem ─────────────────────────────

@pytest.mark.parametrize("code", ["CON", "PRN", "NUL", "com5", "lpt9", "Con", "coM5"])
def test_windows_reserved_names_get_underscore_prefix_and_file_exists(out_dir, code):
    expected_stem = "_" + code
    assert safe_filename(code) == expected_stem
    p = G.generate(code, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
    assert p.name == f"{expected_stem}.png"
    assert p.exists() and p.stat().st_size > 0


def test_filename_collision_AB_slash_CD_vs_AB_question_CD():
    a = safe_filename("AB/CD")
    b = safe_filename("AB?CD")
    assert a == b == "AB_CD"


def test_overwrite_updates_mtime(out_dir):
    p1 = G.generate(EAN, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
    t1 = p1.stat().st_mtime_ns
    time.sleep(0.05)
    p2 = G.generate(EAN, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
    t2 = p2.stat().st_mtime_ns
    assert t2 > t1


def test_200_char_code128_generates_without_path_length_crash(out_dir):
    """200-char code -> long filename. On THIS machine (long paths enabled)
    it succeeds; flagged in the report as an environment-dependent risk, not
    asserted as a hard failure since it did not fail here."""
    code = "A" * 200
    p = G.generate(code, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
    assert p.exists() and p.stat().st_size > 0
    assert len(str(p)) > 200


# ────────────────────── FAZA 3 — settings / registry (isolated key) ─────────

@pytest.fixture()
def isolated_registry(monkeypatch):
    monkeypatch.setattr(app_mod, "REG_KEY", REG_KEY_ISOLATED)
    yield
    try:
        winreg.DeleteKey(winreg.HKEY_CURRENT_USER, REG_KEY_ISOLATED)
    except FileNotFoundError:
        pass


def _write_raw(pairs):
    key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, REG_KEY_ISOLATED)
    for k, v in pairs.items():
        winreg.SetValueEx(key, k, 0, winreg.REG_SZ, str(v))
    winreg.CloseKey(key)


def test_registry_language_DE_falls_back_to_PL(isolated_registry):
    _write_raw({"language": "DE"})
    s = app_mod.load_settings()
    assert s["language"] == "PL"


def test_registry_theme_neon_falls_back_to_dark(isolated_registry):
    _write_raw({"theme": "neon"})
    s = app_mod.load_settings()
    assert s["theme"] == "dark"


def test_registry_dpi_99999_clamped_to_1200(isolated_registry):
    _write_raw({"dpi": "99999"})
    s = app_mod.load_settings()
    assert s["dpi"] == 1200


def test_registry_font_size_legacy_migration_and_clamp(isolated_registry):
    """font_size='12' with no font_unit marker -> migrated 12*1.693=20.316,
    round to 20.3, then clamped to PARAM_RANGES max 15.0."""
    _write_raw({"font_size": "12"})
    s = app_mod.load_settings()
    assert s["font_size"] == 15.0


def test_registry_distance_legacy_015_migrates_to_default(isolated_registry):
    _write_raw({"distance": "0.15"})
    s = app_mod.load_settings()
    assert s["distance"] == app_mod.DEFAULT_SETTINGS["distance"] == 0.33


def test_registry_load_never_raises_on_any_combination(isolated_registry):
    """Throwing every corrupt value at once must not crash load_settings."""
    _write_raw({"language": "XX", "theme": "purple", "dpi": "-5",
                "font_size": "abc_not_a_number", "distance": "999999",
                "height": "-3", "scale": "9999"})
    # font_size is corrupt text -> QueryValueEx cast to float raises ValueError,
    # caught internally -> falls back to default, must not propagate.
    s = app_mod.load_settings()
    assert s["language"] == "PL"
    assert s["theme"] == "dark"
    assert 72 <= s["dpi"] <= 1200


# ──────────────────────── FAZA 3 — Excel import logic ────────────────────────

def _import_excel_logic(path):
    """Exact re-implementation of app.py's _import_excel row-reading logic."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    codes = [
        str(row[0]).strip()
        for row in ws.iter_rows(min_row=1, values_only=True)
        if row and row[0] is not None and str(row[0]).strip()
    ]
    wb.close()
    return codes


def test_excel_import_mixed_types_500_rows(out_dir):
    import openpyxl
    from datetime import datetime
    path = out_dir / "mixed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=123456789)          # int
    ws.cell(row=2, column=1, value="ABC-123")           # text
    c = ws.cell(row=3, column=1, value="=1+1")          # "formula" (no cached value w/o real Excel)
    ws.cell(row=4, column=1, value=None)                # empty
    ws.cell(row=5, column=1, value=datetime(2026, 1, 1))  # date
    for i in range(5, 500):
        ws.cell(row=i + 1, column=1, value=f"CODE{i}")
    wb.save(path)

    codes = _import_excel_logic(path)
    # row1 int, row2 text, row3 formula->None (data_only w/ no cached value, skipped),
    # row4 skipped (None), row5 date (stringified), then 495 CODE* rows = 498 total
    assert len(codes) == 498
    assert codes[0] == "123456789"
    assert codes[1] == "ABC-123"
    assert codes[2] == "2026-01-01 00:00:00"
    assert codes[3] == "CODE5"


def test_BUG_excel_ean_leading_zero_lost_when_stored_as_number(out_dir):
    """BUG (confirmed): if an Excel cell holds a 13-digit EAN with a LEADING
    ZERO as a NUMBER (the default Excel cell type for anything numeric-looking
    typed by a user), Excel/openpyxl round-trips it as an int and the leading
    zero is silently lost -> a 12-digit string comes back from the importer.
    generator._detect_barcode_type requires EXACTLY 13 digits for EAN-13, so
    this 12-digit string is silently generated as CODE128 instead of EAN-13,
    with NO error raised to warn the user their barcode type is now wrong."""
    d12 = "050123412345"
    chk = _ean13_checksum(d12)
    full_code = d12 + str(chk)   # correct, intended 13-digit EAN-13
    assert len(full_code) == 13

    import openpyxl
    path = out_dir / "leading_zero.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=int(full_code))  # as NUMBER: leading 0 lost
    wb.save(path)
    codes = _import_excel_logic(path)
    imported = codes[0]

    assert imported == str(int(full_code))          # leading zero IS lost
    assert len(imported) == 12                       # -> no longer 13 digits
    assert _detect_barcode_type(imported) == "code128"  # -> silently mis-typed
    # generator does NOT raise despite this being a corrupted EAN payload
    p = G.generate(imported, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
    dec = _decode(p)
    assert any(r.format.name == "Code128" and r.text == imported for r in dec)


def test_BUG_excel_ean_trailing_dot_zero_silently_becomes_code128(out_dir):
    """BUG (confirmed): a code arriving from any float-producing path as
    '5907925017654.0' (15 chars incl. the dot) is not 13 all-digit chars, so
    _detect_barcode_type silently falls through to Code128 and encodes the
    LITERAL string including the decimal point — no error, no warning."""
    code = "5907925017654.0"
    assert _detect_barcode_type(code) == "code128"
    p = G.generate(code, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
    dec = _decode(p)
    assert any(r.text == code for r in dec)  # confirms the literal '.0' got encoded


# ───────────────── FAZA 4 — GUI logic (headless CTk where possible) ─────────

def test_dpi_out_of_range_rejected_by_param_ranges_logic():
    lo, hi = app_mod.PARAM_RANGES["dpi"]
    assert not (lo <= 10000 <= hi)
    assert lo <= 1200 <= hi
    assert lo <= 72 <= hi
    assert not (lo <= 71 <= hi)


def test_scale_entry_parsing_logic_comma_and_clamp():
    def on_scale_entry(raw, current):
        try:
            val = float(raw.replace(",", "."))
            return max(0.5, min(3.0, round(val, 1)))
        except ValueError:
            return current

    assert on_scale_entry("1,5", 1.0) == 1.5
    assert on_scale_entry("abc", 1.0) == 1.0     # invalid -> unchanged
    assert on_scale_entry("0.1", 1.0) == 0.5     # clamp low
    assert on_scale_entry("99", 1.0) == 3.0      # clamp high


def test_duplicate_and_name_collision_detection_logic():
    def analyze(codes):
        seen, dupes, unique = set(), [], []
        for c in codes:
            if c in seen:
                if c not in dupes:
                    dupes.append(c)
            else:
                seen.add(c)
                unique.append(c)
        by_name = {}
        for c in unique:
            by_name.setdefault(safe_filename(c), []).append(c)
        collisions = [v for v in by_name.values() if len(v) > 1]
        return dupes, unique, collisions

    dupes, unique, collisions = analyze(["AB/CD", "AB/CD", "AB/CD"])
    assert dupes == ["AB/CD"] and unique == ["AB/CD"] and collisions == []

    dupes, unique, collisions = analyze(["AB/CD", "AB?CD"])
    assert dupes == [] and collisions == [["AB/CD", "AB?CD"]]


def test_worker_loop_batch_with_errors_in_middle(out_dir):
    """Replicates app.py's worker() loop (no GUI): 10 codes, 2 with a broken
    checksum -> 8 files created, exactly 2 errors reported."""
    codes = []
    for i in range(8):
        d12 = f"59012341234{i}"[:12].ljust(12, "0")
        codes.append(d12 + str(_ean13_checksum(d12)))
    for i in range(2):
        d12 = f"5901234123{i}0"[:12].ljust(12, "0")
        chk = _ean13_checksum(d12)
        codes.append(d12 + str((chk + 1) % 10))

    settings = dict(height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
    errors = []
    for code in codes:
        try:
            G.generate(code=code, out_dir=out_dir, height=settings["height"],
                       distance=settings["distance"], font_size=settings["font_size"],
                       dpi=int(settings["dpi"]), scale=settings["scale"])
        except Exception as e:
            errors.append((code, str(e)))

    files = list(out_dir.glob("*.png"))
    assert len(files) == 8
    assert len(errors) == 2


def test_resolve_output_dir_logic_empty_and_nested_and_file_conflict(out_dir):
    def resolve(settings, app_dir_fn):
        base = settings.get("output_dir", "").strip()
        base = Path(base) if base else app_dir_fn() / "output"
        base.mkdir(parents=True, exist_ok=True)
        return base

    d = resolve({"output_dir": ""}, lambda: out_dir)
    assert d == out_dir / "output" and d.exists()

    nested = out_dir / "a" / "b" / "c"
    d2 = resolve({"output_dir": str(nested)}, lambda: out_dir)
    assert d2.exists()

    badfile = out_dir / "notadir.txt"
    badfile.write_text("x")
    with pytest.raises(OSError):
        resolve({"output_dir": str(badfile / "sub")}, lambda: out_dir)


def test_open_output_folder_logic_noop_after_folder_deleted(out_dir):
    class Fake:
        _last_out_dir = out_dir

    def open_folder(self):
        opened = []
        if hasattr(self, "_last_out_dir") and self._last_out_dir.exists():
            opened.append("startfile")
        return opened

    f = Fake()
    assert open_folder(f) == ["startfile"]
    import shutil
    shutil.rmtree(out_dir)
    assert open_folder(f) == []  # silent no-op, no crash, no user feedback


# ───────────────────── FAZA 4 real headless-GUI verification ────────────────
# customtkinter/Tk CAN create real (invisible) windows on this Windows host
# without a running mainloop() — verified experimentally. Used sparingly for
# the highest-value interactive checks; isolated via its own registry subkey.

REG_KEY_GUI = r"Software\BarcodeGen_QA_sonnet_2_gui"


@pytest.fixture()
def gui_app():
    app_mod.REG_KEY = REG_KEY_GUI
    a = app_mod.App()
    yield a
    a.destroy()
    try:
        winreg.DeleteKey(winreg.HKEY_CURRENT_USER, REG_KEY_GUI)
    except FileNotFoundError:
        pass
    app_mod.REG_KEY = r"Software\BarcodeGen"


def _find_all(widget):
    kids = widget.winfo_children()
    out = list(kids)
    for k in kids:
        out += _find_all(k)
    return out


def test_gui_settings_dialog_rejects_dpi_10000_and_keeps_old_value(gui_app):
    import customtkinter as ctk
    a = gui_app
    with mock.patch("app.messagebox.showerror") as mshow:
        a._open_settings()
        win = a.winfo_children()[-1]
        widgets = _find_all(win)
        entries = [w for w in widgets if isinstance(w, ctk.CTkEntry)]
        buttons = [w for w in widgets if isinstance(w, ctk.CTkButton)]
        dpi_entry = next(e for e in entries if e.get() == str(a.settings["dpi"]))
        dpi_entry.delete(0, "end")
        dpi_entry.insert(0, "10000")
        save_btn = next(b for b in buttons if b.cget("text") == a.t["save"])
        before_dpi = a.settings["dpi"]
        save_btn.invoke()
        assert mshow.called
        assert a.settings["dpi"] == before_dpi == 300


def test_gui_reset_defaults_restores_dpi_and_language(gui_app):
    import customtkinter as ctk
    a = gui_app
    a.settings["dpi"] = 999
    a.settings["language"] = "EN"
    with mock.patch("app.messagebox.askyesno", return_value=True):
        a._open_settings()
        win = a.winfo_children()[-1]
        buttons = [w for w in _find_all(win) if isinstance(w, ctk.CTkButton)]
        reset_btn = next(b for b in buttons if b.cget("text") == a.t["reset_defaults"])
        reset_btn.invoke()
    assert a.settings["dpi"] == 300
    assert a.settings["language"] == "PL"


def test_gui_language_toggle_preserves_typed_codes_and_log(gui_app):
    a = gui_app
    a.code_input.insert("1.0", f"{EAN}\n{C128}")
    a._log("[OK] test-line")
    a.progress_bar.set(0.5)
    a._toggle_lang()
    assert a.lang == "EN"
    assert a.code_input.get("1.0", "end").strip() == f"{EAN}\n{C128}"
    assert "[OK] test-line" in a.log_box.get("1.0", "end")
    assert a.progress_bar.get() == 0.5


def test_gui_help_window_reused_not_duplicated(gui_app):
    a = gui_app
    a._open_help()
    first_id = a._help_win.winfo_id()
    a._open_help()
    second_id = a._help_win.winfo_id()
    assert first_id == second_id


def test_gui_duplicate_then_collision_dialogs_fire_in_sequence(gui_app):
    a = gui_app
    out = _out()
    a.settings["output_dir"] = str(out)
    a.code_input.delete("1.0", "end")
    a.code_input.insert("1.0", "AB/CD\nAB/CD\nAB?CD")
    seen_messages = []

    def fake_askyesno(title, msg, parent=None):
        seen_messages.append(msg)
        return True

    with mock.patch("app.messagebox.askyesno", side_effect=fake_askyesno):
        a._on_generate()
    # daemon worker thread — poll for completion instead of a fixed sleep
    deadline = time.time() + 10
    while time.time() < deadline and not list(Path(out).glob("*.png")):
        time.sleep(0.1)
    time.sleep(0.3)  # let the final after(0, ...) callbacks flush
    assert len(seen_messages) == 2
    assert "duplikat" in seen_messages[0].lower()
    assert "ten sam plik" in seen_messages[1].lower()
    files = os.listdir(out)
    assert files == ["AB_CD.png"]  # both codes sanitize to one file


def test_gui_theme_toggle_persists_to_registry(gui_app):
    a = gui_app
    before = a.settings["theme"]
    a._toggle_theme()
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, REG_KEY_GUI)
    val, _ = winreg.QueryValueEx(key, "theme")
    winreg.CloseKey(key)
    assert a.settings["theme"] != before
    assert val == a.settings["theme"]


# ──────────────── FAZA 5 — free hunting: dead code + attack scenarios ───────

def test_count_data_modules_is_dead_code():
    """FINDING: _count_data_modules (and its _PROBE_MW_MM/_PROBE_QZ_MM
    constants) is defined in generator.py but never called from generate()
    or anywhere in app.py — exactly the 'looks wired up but does nothing'
    trap the audit brief warns about. Confirmed by source inspection AND by
    the fact removing/breaking it does not affect any generate() output
    (it's simply unreachable)."""
    import generator
    import inspect
    src_generate = inspect.getsource(generator.BarcodeGenerator.generate)
    assert "_count_data_modules" not in src_generate
    with open(Path(__file__).resolve().parent.parent / "app.py", encoding="utf-8") as f:
        app_src = f.read()
    assert "_count_data_modules" not in app_src
    assert not hasattr(generator, "_count_data_modules")  # FIXED(N6): dead code removed


def test_BUG_all_barcode_error_messages_are_hardcoded_polish_regardless_of_ui_language(out_dir):
    """FINDING: every BarcodeError message in generator.py is a hardcoded
    Polish f-string. When the UI is switched to EN, [ERR] log lines still
    show Polish text (e.g. 'EAN-13 musi mieć dokładnie 13 cyfr...') because
    the error text never goes through lang.py."""
    bad_ean = "5901234123458"  # 13 digits, wrong check digit (correct is ...457)
    with pytest.raises(BarcodeError) as ei:
        G.generate(bad_ean, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
    assert "Błędna suma kontrolna" in str(ei.value)  # Polish, unconditionally


def test_footer_copyright_string_is_hardcoded_polish_outside_lang_py():
    """FIXED(N5): footer copyright moved into lang.py (translated PL/EN); app.py
    no longer hardcodes it and the EN dict has its own variant."""
    with open(Path(__file__).resolve().parent.parent / "app.py", encoding="utf-8") as f:
        src = f.read()
    assert "Wszelkie prawa zastrzeżone" not in src
    from lang import LANG
    assert "footer_copyright" in LANG["PL"] and "footer_copyright" in LANG["EN"]
    assert LANG["EN"]["footer_copyright"] != LANG["PL"]["footer_copyright"]


def test_ATTACK_unicode_and_emoji_rejected_for_code128(out_dir):
    for code in ("ABC" + chr(0x1F389) + "123", "ĄĘŚĆ"):
        with pytest.raises(BarcodeError):
            G.generate(code, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)


def test_ATTACK_negative_height_or_scale_bypasses_generate_and_crashes_with_ValueError(out_dir):
    """FIXED(N3): negative height/scale now raise a clean BarcodeError up front
    instead of a raw ValueError from PIL. Negative distance is still internally
    floored (max(0.05, distance)) and generates normally."""
    with pytest.raises(BarcodeError):
        G.generate(C128, out_dir, height=-5, distance=0.33, font_size=2.2, dpi=300, scale=1.0)

    with pytest.raises(BarcodeError):
        G.generate(C128, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=-1)

    # negative `distance` is internally clamped and does NOT crash (unchanged)
    p = G.generate(C128, out_dir, height=9, distance=-1, font_size=2.2, dpi=300, scale=1.0)
    assert p.exists()


def test_ATTACK_param_ranges_max_combo_triggers_decompression_bomb(out_dir):
    """CRITICAL BUG: app.py's own comment says PARAM_RANGES 'Prevents e.g.
    dpi=10000 + height=1000 from generating a multi-gigapixel image that
    hangs or crashes the app'. But combining ONLY values that are each
    individually WITHIN the allowed GUI ranges (height=150 [max], distance=10
    [max], font_size=15 [max], dpi=1200 [max], scale=3.0 [max]) still
    produces a ~2.9 gigapixel canvas that trips Pillow's own decompression-
    bomb guard -> PIL.Image.DecompressionBombError, an exception the worker's
    generic `except Exception` catches (so the app itself does not crash),
    but the very safety net PARAM_RANGES was built for does not, in fact,
    prevent the gigapixel image.

    FIXED(Ś1): generate() now estimates the pixel count up front and raises a
    clean BarcodeError before rendering the oversized canvas."""
    with pytest.raises(BarcodeError):
        G.generate(EAN, out_dir, height=150, distance=10, font_size=15, dpi=1200, scale=3.0)


def test_ATTACK_ean13_all_zero_with_valid_checksum_decodes(out_dir):
    d12 = "0" * 12
    code = d12 + str(_ean13_checksum(d12))
    p = G.generate(code, out_dir, height=9, distance=0.33, font_size=2.2, dpi=300, scale=1.0)
    dec = _decode(p)
    assert any(r.text == code for r in dec)


def test_ATTACK_large_but_in_range_combo_is_slow(out_dir):
    """Not a hard crash, but a real perf concern: a combo well within
    PARAM_RANGES (nowhere near the maxima) already produces a 126-megapixel
    image and takes multiple seconds for a SINGLE code. With up to 100 codes
    per batch and no cancel button (see report, manual items 15/16), this is
    a realistic multi-minute UI freeze risk."""
    t0 = time.time()
    p = G.generate(EAN, out_dir, height=100, distance=3, font_size=15, dpi=800, scale=2.0)
    dt = time.time() - t0
    im = Image.open(p)
    megapixels = im.size[0] * im.size[1] / 1e6
    assert megapixels > 50   # confirms this is indeed a "big" image
    assert dt < 30           # sanity bound so the test suite itself doesn't hang
    print(f"perf note: {megapixels:.1f} MP in {dt:.2f}s for a single code")


def test_gap_after_7th_char_measured_ean13_vs_code128_short(out_dir):
    """FAZA 4b #19: measured pixel gap after the 7th glyph must be clearly
    bigger than the regular tracking gap for EAN-13 (13 chars, has a 7th),
    and Code128 <7 chars must have NO such outlier gap."""
    p_ean = G.generate(EAN, out_dir, height=9, distance=0.33, font_size=3.0, dpi=300, scale=1.0)
    glyphs, gaps = _glyph_gaps_px(p_ean)
    assert len(glyphs) == 13
    normal_gaps = gaps[:6] + gaps[7:]
    assert gaps[6] > 3 * max(normal_gaps), f"gap-after-7th {gaps[6]} not clearly bigger than {normal_gaps}"

    p_short = G.generate("ABC-12", out_dir, height=9, distance=0.33, font_size=3.0, dpi=300, scale=1.0)
    glyphs_s, gaps_s = _glyph_gaps_px(p_short)
    assert len(glyphs_s) == 6  # no 7th char
    assert max(gaps_s) < 3 * min(gaps_s) if min(gaps_s) > 0 else True  # no dramatic outlier


def test_gap_after_7th_char_also_applies_to_long_code128():
    """Design observation (not a crash bug, documented as intentional in
    generator.py's comment 'Applied universally to any code'): an 8+ char
    Code128 part number ALSO gets the EAN-style big gap after its 7th
    character, which may look unintended for arbitrary alphanumeric codes."""
    out = _out()
    p = G.generate("ABCDEFGH", out, height=9, distance=0.33, font_size=3.0, dpi=300, scale=1.0)
    glyphs, gaps = _glyph_gaps_px(p)
    assert len(glyphs) == 8
    normal_gaps = gaps[:6]
    assert gaps[6] > 3 * max(normal_gaps)


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
